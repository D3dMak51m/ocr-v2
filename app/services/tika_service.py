import asyncio
import os
import io
import logging
import tempfile
import zipfile
from typing import List, Tuple
from zipfile import ZipFile
import httpx

import fitz  # PyMuPDF
from PIL import Image

from config import settings
from core import file_types
from core.exceptions import ExternalServiceError, FileProcessingError
from core.schemas import DocOcrResult, ImageOcrResult
from services import image_service, utils

# Initialize logger
logger = logging.getLogger(__name__)


# ==================== PDF Processing ====================

def _are_image_tiles(image_blocks: list, page_width: float, page_height: float) -> bool:
    """
    Проверяет, являются ли image_blocks тайлами одного изображения.
    Если их bbox'ы покрывают >70% площади страницы — это тайлы.
    """
    if len(image_blocks) <= 1:
        return False

    page_area = page_width * page_height
    if page_area <= 0:
        return False

    # Считаем суммарную площадь всех image-блоков
    total_img_area = 0
    for block in image_blocks:
        bbox = block["bbox"]
        w = bbox[2] - bbox[0]
        h = bbox[3] - bbox[1]
        total_img_area += w * h

    coverage = total_img_area / page_area

    if coverage >= 0.70:
        logger.info(f"Detected tiled image: {len(image_blocks)} blocks cover {coverage:.0%} of page")
        return True

    return False


async def _async_process_pdf_hybrid(filepath: str) -> Tuple[str, List[ImageOcrResult]]:
    """
    Умный парсинг PDF: читает блоки страницы сверху вниз.
    Нативный текст копирует, встроенные картинки отправляет в OCR.
    Если картинки разбиты на тайлы, склеивает их в одно изображение.
    Текст из картинок НЕ дублируется в data.text — он доступен в images[].text.
    """
    full_text_parts = []
    all_images_results = []

    logger.info(f"Starting hybrid parsing for PDF: {filepath}")

    try:
        # Открываем PDF
        with fitz.open(filepath) as doc:
            for page_num in range(len(doc)):
                page = doc[page_num]

                # Получаем словарь структуры страницы (текст + картинки)
                page_dict = page.get_text("dict")
                blocks = page_dict.get("blocks", [])
                page_width = page_dict.get("width", page.rect.width)
                page_height = page_dict.get("height", page.rect.height)

                # Сортируем блоки по вертикали (сверху вниз)
                blocks.sort(key=lambda b: b["bbox"][1])

                # Разделяем блоки на текстовые и изображения
                text_blocks = [b for b in blocks if b["type"] == 0]
                image_blocks = [b for b in blocks if b["type"] == 1 and b.get("image")]

                # Обрабатываем текстовые блоки (только нативный цифровой текст)
                for block in text_blocks:
                    text_content = ""
                    for line in block.get("lines", []):
                        for span in line.get("spans", []):
                            text_content += span.get("text", "") + " "
                        text_content += "\n"

                    cleaned_text = text_content.strip()
                    if cleaned_text:
                        full_text_parts.append(cleaned_text)

                # Обрабатываем изображения: проверяем, не тайлы ли это
                if image_blocks:
                    is_tiled = _are_image_tiles(image_blocks, page_width, page_height)

                    if is_tiled:
                        # Тайлы → рендерим всю страницу в высоком разрешении (2x)
                        logger.info(f"Page {page_num + 1}: rendering full page at 2x for {len(image_blocks)} tiles")
                        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                        merged_img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

                        result = await asyncio.to_thread(image_service.process_image_from_pil, merged_img)
                        result.filename = f"page_{page_num + 1}_img.png"
                        all_images_results.append(result)
                    else:
                        # Обрабатываем каждую картинку отдельно (если это не тайлы)
                        for block in image_blocks:
                            image_bytes = block.get("image")
                            img = Image.open(io.BytesIO(image_bytes))
                            if img.mode != "RGB":
                                img = img.convert("RGB")

                            result = await asyncio.to_thread(image_service.process_image_from_pil, img)
                            result.filename = f"page_{page_num + 1}_img.png"
                            all_images_results.append(result)

        # Склеиваем только нативный текст (без OCR-текста из картинок)
        combined_text = "\n\n".join(full_text_parts)
        return combined_text, all_images_results

    except Exception as e:
        logger.error(f"Hybrid PDF processing failed: {e}")
        raise FileProcessingError(f"Hybrid PDF processing failed: {e}")


# ==================== Tika Integration ====================

def _has_encoding_issues(text: str) -> bool:
    """
    Проверяет текст на проблемы кодировки:
    - наличие Unicode replacement character (U+FFFD)
    - >10% символов в диапазоне 127-160 (Latin-1 control chars)
    """
    if not text:
        return False

    if '\ufffd' in text:
        return True

    special_char_count = sum(1 for c in text if 127 < ord(c) < 160)
    if len(text) > 100 and special_char_count / len(text) > 0.1:
        return True

    return False


async def _async_tika_get_text(filepath: str) -> str:
    """
    Извлекает текст из документа через Apache Tika.
    Включает proper encoding handling и проверку на encoding issues.
    """
    tika_url = f"{settings.TIKA_SERVER_URL}/tika"
    logger.info(f"Sending file to Tika for text extraction: {tika_url}")

    try:
        with open(filepath, "rb") as f:
            file_data = f.read()

        async with httpx.AsyncClient(timeout=120.0) as client:
            response = await client.put(
                tika_url, content=file_data,
                headers={
                    "Accept": "text/plain; charset=utf-8",
                    "Content-Type": "application/octet-stream",
                    "Accept-Charset": "utf-8",
                }
            )
            response.raise_for_status()

            # Принудительно указываем UTF-8 перед декодированием
            response.encoding = "utf-8"
            try:
                text = response.text
            except UnicodeDecodeError:
                import chardet
                detected = chardet.detect(response.content)
                encoding = detected.get('encoding', 'utf-8')
                logger.info(f"Detected encoding: {encoding}")
                text = response.content.decode(encoding, errors='replace')

        # Проверяем качество извлечённого текста
        if _has_encoding_issues(text):
            logger.warning(f"Encoding issues detected in Tika output for {filepath}")

        return text

    except httpx.HTTPStatusError as e:
        msg = f"Tika server returned error: {e.response.status_code}"
        logger.error(msg)
        raise ExternalServiceError("Tika", msg)
    except Exception as e:
        msg = f"Tika text extraction failed: {e}"
        logger.error(msg)
        raise ExternalServiceError("Tika", msg)


async def _async_tika_extract_embedded_files(filepath: str) -> List[ImageOcrResult]:
    """
    Асинхронно распаковывает встроенные картинки через Tika /unpack.
    Возвращает список ImageOcrResult для каждого найденного изображения.
    """
    tika_url = f"{settings.TIKA_SERVER_URL}/unpack"
    logger.info(f"Sending file to Tika for unpacking embedded files: {tika_url}")
    ocr_results = []

    try:
        with open(filepath, "rb") as f:
            file_data = f.read()

        async with httpx.AsyncClient(timeout=120.0) as client:
            response = await client.put(
                tika_url,
                content=file_data,
                headers={"Accept": "application/zip"}
            )

            # Handle 204 No Content (no embedded files)
            if response.status_code == 204:
                logger.info(f"No embedded files found in {filepath} (HTTP 204)")
                return ocr_results

            # Проверяем HTTP-ошибки (перенесено из sync версии)
            response.raise_for_status()

            if not response.content:
                logger.info(f"Empty response content from Tika for {filepath}")
                return ocr_results

            # Проверяем magic bytes ZIP (PK)
            if len(response.content) < 4 or response.content[:2] != b'PK':
                logger.info(f"Response from Tika is not a ZIP file for {filepath}")
                return ocr_results

        logger.info(f"Successfully received embedded files from Tika for {filepath}")

        with tempfile.NamedTemporaryFile(suffix=".zip", delete=True) as tmpzip:
            tmpzip.write(response.content)
            tmpzip.seek(0)

            try:
                with zipfile.ZipFile(tmpzip, "r") as zip_ref:
                    for item_name in zip_ref.namelist():
                        if item_name.startswith('__'):
                            continue
                        try:
                            file_bytes = zip_ref.read(item_name)
                            result = await asyncio.to_thread(
                                image_service.process_image_from_bytes, file_bytes
                            )
                            result.filename = os.path.basename(item_name)
                            ocr_results.append(result)
                        except Exception as e:
                            logger.warning(f"Could not process embedded file '{item_name}': {e}")
            except zipfile.BadZipFile as e:
                logger.warning(f"Invalid ZIP file returned from Tika for {filepath}: {e}")

    except httpx.HTTPStatusError as e:
        msg = f"Tika server returned error during unpacking: {e.response.status_code}"
        logger.error(msg)
        raise ExternalServiceError("Tika", msg)
    except ExternalServiceError:
        raise
    except Exception as e:
        logger.error(f"Tika unpacking failed: {e}")

    return ocr_results


# ==================== Excel Processing ====================

def _is_excel_file(filepath: str) -> bool:
    """Check if a file is an Excel file based on extension."""
    ext = os.path.splitext(filepath)[1].lower()
    return ext in ['.xls', '.xlsx', '.xlsm', '.xlsb']


def _extract_excel_text(filepath: str) -> str:
    """
    Извлекает текст из Excel файлов с правильной поддержкой кодировки.
    Использует openpyxl (XLSX) или xlrd (XLS) напрямую вместо Tika.
    Fallback на pandas если основные библиотеки недоступны.
    """
    ext = os.path.splitext(filepath)[1].lower()

    # 1. XLSX через openpyxl (primary)
    if ext in ['.xlsx', '.xlsm', '.xlsb']:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filepath, data_only=True, read_only=True)
            all_text = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                all_text.append(f"Sheet: {sheet_name}\n")

                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                    if row_text:
                        all_text.append('\t'.join(row_text))

            wb.close()
            result = '\n'.join(all_text)
            if result.strip():
                logger.info(f"Extracted Excel text via openpyxl ({len(result)} chars)")
                return result

        except ImportError:
            logger.warning("openpyxl not installed, falling back")
        except Exception as e:
            logger.warning(f"openpyxl extraction failed: {e}")

    # 2. XLS через xlrd
    if ext == '.xls':
        try:
            import xlrd

            book = xlrd.open_workbook(filepath, encoding_override=None)
            all_text = []

            for sheet_idx in range(book.nsheets):
                sheet = book.sheet_by_index(sheet_idx)
                all_text.append(f"Sheet: {sheet.name}\n")

                for row_idx in range(sheet.nrows):
                    row_text = []
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        if cell.value:
                            row_text.append(str(cell.value))
                    if row_text:
                        all_text.append('\t'.join(row_text))

            result = '\n'.join(all_text)
            if result.strip():
                logger.info(f"Extracted Excel text via xlrd ({len(result)} chars)")
                return result

        except ImportError:
            logger.warning("xlrd not installed for XLS processing")
        except Exception as e:
            logger.warning(f"xlrd extraction failed: {e}")

    # 3. Pandas — last resort
    try:
        import pandas as pd

        engine = 'xlrd' if ext == '.xls' else 'openpyxl'
        df_dict = pd.read_excel(filepath, sheet_name=None, engine=engine)

        all_text = []
        for sheet_name, df in df_dict.items():
            all_text.append(f"Sheet: {sheet_name}\n")
            all_text.append(df.to_string())

        result = '\n'.join(all_text)
        if result.strip():
            logger.info(f"Extracted Excel text via pandas ({len(result)} chars)")
            return result

    except ImportError:
        logger.warning("pandas not installed for Excel processing")
    except Exception as e:
        logger.warning(f"pandas extraction failed: {e}")

    return ""


def _extract_images_from_xlsx(filepath: str) -> List[ImageOcrResult]:
    """
    Extract images from XLSX files directly.
    XLSX files are actually ZIP archives with a specific structure.
    """
    ocr_results = []
    logger.info(f"Attempting to extract images from XLSX file: {filepath}")

    try:
        with ZipFile(filepath, 'r') as zip_file:
            all_files = zip_file.namelist()

            # Images in XLSX are typically stored in xl/media/ directory
            image_files = [f for f in all_files if f.startswith('xl/media/')]

            if not image_files:
                logger.info(f"No images found in XLSX file: {filepath}")
                return ocr_results

            logger.info(f"Found {len(image_files)} images in XLSX file")

            for image_file in image_files:
                try:
                    image_bytes = zip_file.read(image_file)
                    result = image_service.process_image_from_bytes(image_bytes)
                    result.filename = os.path.basename(image_file)
                    ocr_results.append(result)
                    logger.info(f"Successfully processed image: {image_file}")
                except Exception as e:
                    logger.warning(f"Could not process image '{image_file}': {e}")

    except zipfile.BadZipFile:
        logger.warning(f"File {filepath} is not a valid ZIP/XLSX file")
    except Exception as e:
        logger.error(f"Error extracting images from XLSX: {e}")

    return ocr_results


# ==================== Main Router ====================

async def process_document_with_tika(filepath: str, file_type: str) -> DocOcrResult:
    """Главная асинхронная функция маршрутизации документа."""

    image_results = []
    formatted_text = ""

    # Если это PDF, используем наш умный гибридный парсер
    if file_type == file_types.TYPE_PDF:
        raw_text, image_results = await _async_process_pdf_hybrid(filepath)
        formatted_text = utils.text_formatting(raw_text)

        return DocOcrResult(
            text=formatted_text,
            images=image_results,
            service="hybrid_fitz_paddle",
        )

    # Для Excel — используем прямое чтение (лучше для кириллицы/UZ)
    if _is_excel_file(filepath):
        raw_text = await asyncio.to_thread(_extract_excel_text, filepath)

        # Fallback на Tika если прямое чтение не дало результата
        if not raw_text.strip():
            logger.info("Direct Excel extraction returned empty, falling back to Tika")
            raw_text = await _async_tika_get_text(filepath)

        # Извлекаем картинки из XLSX
        image_results = await asyncio.to_thread(_extract_images_from_xlsx, filepath)
        if not image_results:
            image_results = await _async_tika_extract_embedded_files(filepath)

        formatted_text = utils.text_formatting(raw_text)

        return DocOcrResult(
            text=formatted_text,
            images=image_results,
            service="excel_paddle",
        )

    # Для остальных форматов (Word, PPT, RTF, txt) используем Tika
    raw_text = await _async_tika_get_text(filepath)
    image_results = await _async_tika_extract_embedded_files(filepath)
    formatted_text = utils.text_formatting(raw_text)

    return DocOcrResult(
        text=formatted_text,
        images=image_results,
        service="tika_paddle",
    )
